#!/usr/bin/env python3
"""
Exports home_cost_basis_ledger.csv into a formatted Excel workbook with tabs:
Summary, Capital Improvements, Needs Review, Repairs & Maintenance (excluded),
Personal Property (excluded).

Usage:
    python export_review_workbook.py --ledger path/to/ledger.csv --out path/to/output.xlsx
"""
import argparse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT = Font(name="Arial")
MONEY_FMT = '$#,##0.00;($#,##0.00);"-"'

TABS = {
    "Capital Improvement": "Capital Improvements",
    "Needs Review": "Needs Review",
    "Repair/Maintenance (excluded)": "Repairs & Maintenance",
    "Excluded - Personal Property": "Personal Property (excl.)",
}

COLS = ["Date", "Source", "Vendor (raw)", "Amount", "Project/Category",
        "Confirmed by Samantha", "Notes"]


def style_sheet(ws, rows):
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(c, "") for c in COLS])
    for i, col in enumerate(COLS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 14 if col != "Vendor (raw)" else 32
        if col == "Notes":
            ws.column_dimensions[letter].width = 30
    amt_col = COLS.index("Amount") + 1
    for row_idx in range(2, ws.max_row + 1):
        c = ws.cell(row=row_idx, column=amt_col)
        c.number_format = MONEY_FMT
        for col_idx in range(1, len(COLS) + 1):
            ws.cell(row=row_idx, column=col_idx).font = BODY_FONT
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ledger", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.ledger, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    grouped = {k: [] for k in TABS}
    for r in rows:
        cls = r.get("Classification", "")
        if cls in grouped:
            grouped[cls].append(r)

    summary.append(["Home Cost Basis Ledger - Summary"])
    summary["A1"].font = Font(bold=True, size=14, name="Arial")
    summary.append([])
    summary.append(["Category", "# Transactions", "Total Amount"])
    for cell in summary[3]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_i = 4
    amt_col_letter = get_column_letter(COLS.index("Amount") + 1)
    tab_row = {}
    for cls, label in TABS.items():
        n = len(grouped[cls])
        formula = f"=SUM('{label}'!{amt_col_letter}2:{amt_col_letter}{n + 1})" if n else 0
        summary.append([label, n, formula])
        summary.cell(row=row_i, column=3).number_format = MONEY_FMT
        tab_row[cls] = row_i
        row_i += 1

    summary.append([])
    ci_ref = f"C{tab_row['Capital Improvement']}"
    nr_ref = f"C{tab_row['Needs Review']}"
    summary.append(["Confirmed Capital Improvement Total (adds to home basis)", "", f"={ci_ref}"])
    summary.cell(row=summary.max_row, column=3).number_format = MONEY_FMT
    summary.cell(row=summary.max_row, column=1).font = Font(bold=True, name="Arial")
    summary.append(["Needs Review — resolve these before finalizing basis", "", f"={nr_ref}"])
    summary.cell(row=summary.max_row, column=3).number_format = MONEY_FMT
    for i, w in enumerate([55, 18, 20], start=1):
        summary.column_dimensions[get_column_letter(i)].width = w

    for cls, label in TABS.items():
        ws = wb.create_sheet(label)
        style_sheet(ws, grouped[cls])

    wb.save(args.out)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
